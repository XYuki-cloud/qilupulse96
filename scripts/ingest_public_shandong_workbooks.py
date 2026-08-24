"""Validate and ingest the public Shandong research workbooks.

The command reads only the four files in ``data/public`` (or an explicitly
provided equivalent directory).  It never calls a weather service and writes
canonical parquet files only below the ignored runtime root.
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
import hashlib
import json
from pathlib import Path
import re
import zipfile
import xml.etree.ElementTree as ET

import openpyxl
import pandas as pd

from da_forecast.sources.manual_realtime_xlsx import read_manual_realtime_prices
from da_forecast.sources.shandong_market_xlsx import (
    merge_identical,
    read_day_ahead_prices,
    read_workbook,
)


PUBLIC_MARKET_FILES = (
    "shandong_market_2024_public.xlsx",
    "shandong_market_2025_public.xlsx",
    "shandong_market_2026-01-01_2026-08-15_public.xlsx",
)
PUBLIC_MANUAL_FILE = "manual_realtime_prices_2026-08-13_2026-08-22_public.xlsx"
PUBLIC_XLSX_FILES = (*PUBLIC_MARKET_FILES, PUBLIC_MANUAL_FILE)
MANIFEST_NAME = "MANIFEST.json"
SCHEMA_VERSION = "public-research-data-v1"
KNOWN_WARNINGS = [
    "The manual workbook contains a duplicate normalized 2026-08-21 00:00 slot; the public parser keeps the first row and records a warning.",
]

MARKET_SHEETS = {
    "实时出清数据": ("目标日期", "时刻", "实时出清电价"),
    "日前出清数据": ("目标日期", "时刻", "日前出清电价"),
    "日前披露数据": (
        "当前日期",
        "目标日期",
        "时刻",
        "相隔天数",
        "负荷信息预测",
        "日前风电总加（MW）",
        "日前光伏总加（MW）",
        "联络线信息预测",
        "竞价空间预测(MW)",
        "负荷率预测",
        "日前地方电厂发电总加（MW）",
        "日前自备机组总加（MW）",
    ),
}
MANUAL_SHEETS = {"Sheet1": ("日期", "时间", "实时电价")}

_FORBIDDEN_TEXT = (
    re.compile(r"(?i)(?<![A-Za-z0-9])(?:[A-Z]:[\\/]|\\\\[^\\/\s]+[\\/])"),
    re.compile(r"(?i)repo_eu_forecast|entso[-_ ]?e|energinet|johannesbroens"),
    re.compile(r"(?i)one1\s+second1"),
    re.compile(r"(?i)\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b"),
    re.compile(r"(?<![\d.])1[3-9]\d{9}(?![\d.])"),
    re.compile(r"(?<![\d.])[1-9]\d{5}(?:18|19|20)\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}(?![\d.])"),
    re.compile(
        r"(?i)(?:AKIA[0-9A-Z]{16}|ghp_[A-Za-z0-9]{20,}|github_pat_[A-Za-z0-9_]{20,}|"
        r"sk-[A-Za-z0-9]{20,}|xox[baprs]-[A-Za-z0-9-]{20,}|-----BEGIN (?:RSA|EC|OPENSSH|PRIVATE) KEY-----)"
    ),
)
_UNSAFE_ZIP_PARTS = ("comments", "vml", "embeddings", "externallink", "custom.xml")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _normalized_header(value: object) -> str:
    return str(value or "").strip().replace(" ", "")


def _text_scan(name: str, content: bytes) -> None:
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        return
    for pattern in _FORBIDDEN_TEXT:
        if pattern.search(text):
            raise ValueError(f"public workbook contains forbidden text marker in {name}")


def _validate_zip_boundary(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        unsafe = [
            name
            for name in names
            if any(part in name.lower() for part in _UNSAFE_ZIP_PARTS)
        ]
        if unsafe:
            raise ValueError(f"{path.name}: workbook contains unsafe package parts: {unsafe[:3]}")
        for name in names:
            if name.endswith(".xml") or name.endswith(".rels"):
                # ``docProps/app.xml`` contains the generic application name
                # emitted by the spreadsheet library, not a workbook author
                # or source identity.  It is checked for package safety below
                # but excluded from personal-text scanning.
                if name != "docProps/app.xml":
                    _text_scan(name, archive.read(name))
        if "docProps/core.xml" not in names:
            raise ValueError(f"{path.name}: missing sanitized core properties")
        core = ET.fromstring(archive.read("docProps/core.xml"))
        identity_names = {
            "{http://purl.org/dc/elements/1.1/}creator",
            "{http://schemas.openxmlformats.org/package/2006/metadata/core-properties}lastModifiedBy",
            "{http://purl.org/dc/elements/1.1/}title",
            "{http://purl.org/dc/elements/1.1/}subject",
            "{http://purl.org/dc/elements/1.1/}description",
        }
        for element in core.iter():
            if element.tag in identity_names and (element.text or "").strip():
                raise ValueError(f"{path.name}: non-empty identity metadata remains")
        for name in names:
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                text = archive.read(name).decode("utf-8", errors="ignore")
                if re.search(r"<f(?:\s|>)|<hyperlink(?:\s|>)", text, flags=re.IGNORECASE):
                    raise ValueError(f"{path.name}: formulas or hyperlinks are not allowed")
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8", errors="ignore")
        if re.search(r'\bstate="(?:hidden|veryHidden)"', workbook_xml, flags=re.IGNORECASE):
            raise ValueError(f"{path.name}: hidden worksheets are not allowed")


def _inspect_workbook(path: Path, schema: dict[str, tuple[str, ...]]) -> dict[str, object]:
    _validate_zip_boundary(path)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        actual_names = tuple(workbook.sheetnames)
        expected_names = tuple(schema)
        if actual_names != expected_names:
            raise ValueError(f"{path.name}: expected sheets {expected_names}, got {actual_names}")
        summaries: list[dict[str, object]] = []
        coverage_values: list[pd.Timestamp] = []
        coverage_sheet = expected_names[0]
        coverage_field = "目标日期" if "目标日期" in schema[coverage_sheet] else "日期"
        for sheet_name, expected_fields in schema.items():
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            try:
                header = tuple(next(rows))
            except StopIteration as exc:
                raise ValueError(f"{path.name}: empty sheet {sheet_name}") from exc
            fields = tuple(str(value or "") for value in header)
            if tuple(_normalized_header(value) for value in fields) != tuple(
                _normalized_header(value) for value in expected_fields
            ):
                raise ValueError(f"{path.name}/{sheet_name}: unexpected fields {fields}")
            data_rows = 0
            for row in rows:
                if all(value is None for value in row):
                    continue
                data_rows += 1
                if sheet_name == coverage_sheet:
                    value = row[expected_fields.index(coverage_field)]
                    parsed = pd.to_datetime(value, errors="coerce")
                    if pd.notna(parsed):
                        coverage_values.append(pd.Timestamp(parsed).normalize())
            summaries.append(
                {
                    "name": sheet_name,
                    "rows": data_rows + 1,
                    "columns": len(fields),
                    "fields": list(fields),
                }
            )
        if not coverage_values:
            raise ValueError(f"{path.name}: no valid date values found")
        return {
            "file": path.name,
            "coverage": {
                "start": min(coverage_values).strftime("%Y-%m-%d"),
                "end": max(coverage_values).strftime("%Y-%m-%d"),
            },
            "sheets": summaries,
        }
    finally:
        workbook.close()


def _schema_for_file(name: str) -> dict[str, tuple[str, ...]]:
    return MANUAL_SHEETS if name == PUBLIC_MANUAL_FILE else MARKET_SHEETS


def _discover(input_dir: Path, *, require_manifest: bool = True) -> dict[str, Path]:
    input_dir = input_dir.expanduser().resolve()
    if not input_dir.is_dir():
        raise FileNotFoundError(f"public input directory not found: {input_dir}")
    actual_xlsx = {path.name for path in input_dir.glob("*.xlsx")}
    expected_xlsx = set(PUBLIC_XLSX_FILES)
    if actual_xlsx != expected_xlsx:
        raise ValueError(
            f"public input directory must contain exactly {sorted(expected_xlsx)}, got {sorted(actual_xlsx)}"
        )
    missing = [name for name in PUBLIC_XLSX_FILES if not (input_dir / name).is_file()]
    if missing:
        raise FileNotFoundError(f"missing public workbook(s): {missing}")
    manifest_path = input_dir / MANIFEST_NAME
    if require_manifest and not manifest_path.is_file():
        raise FileNotFoundError(f"missing {MANIFEST_NAME} in {input_dir}")
    return {name: input_dir / name for name in PUBLIC_XLSX_FILES}


def build_public_manifest(input_dir: str | Path) -> dict[str, object]:
    """Build a manifest from a four-file package without private source paths."""
    input_dir = Path(input_dir).expanduser().resolve()
    files: list[dict[str, object]] = []
    for name, path in _discover(input_dir, require_manifest=False).items():
        inspected = _inspect_workbook(path, _schema_for_file(name))
        files.append(
            {
                **inspected,
                "bytes": path.stat().st_size,
                "sha256": _sha256(path),
            }
        )
    return {
        "schema_version": SCHEMA_VERSION,
        "purpose": "Authorized real-data research inputs for QiluPulse-96; not a production service or benchmark.",
        "generated_at_utc": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source_files_are_private": True,
        "original_workbooks_are_not_modified": True,
        "excluded_source_sheet": "实际披露数据",
        "known_warnings": KNOWN_WARNINGS,
        "files": files,
        "restrictions": [
            "The public copies are derived minimal-field workbooks, not the original source workbooks.",
            "The data is separate from the Apache-2.0 code license.",
            "The data is for research and reproducibility only; no production, trading, or revenue claim is made.",
        ],
    }


def _validate_manifest(input_dir: Path, paths: dict[str, Path]) -> dict[str, object]:
    try:
        manifest = json.loads((input_dir / MANIFEST_NAME).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"invalid {MANIFEST_NAME}") from exc
    if manifest.get("schema_version") != SCHEMA_VERSION:
        raise ValueError(f"unsupported public data schema: {manifest.get('schema_version')!r}")
    if manifest.get("excluded_source_sheet") != "实际披露数据":
        raise ValueError(f"{MANIFEST_NAME}: excluded source sheet marker is missing")
    if manifest.get("known_warnings") != KNOWN_WARNINGS:
        raise ValueError(f"{MANIFEST_NAME}: known warning list differs from the importer contract")
    entries = manifest.get("files")
    if not isinstance(entries, list):
        raise ValueError(f"{MANIFEST_NAME}: files must be a list")
    by_name = {entry.get("file"): entry for entry in entries if isinstance(entry, dict)}
    if set(by_name) != set(PUBLIC_XLSX_FILES):
        raise ValueError(f"{MANIFEST_NAME}: file list does not match the four public workbooks")
    for name, path in paths.items():
        expected = by_name[name]
        inspected = _inspect_workbook(path, _schema_for_file(name))
        if expected.get("bytes") != path.stat().st_size:
            raise ValueError(f"{name}: byte count differs from {MANIFEST_NAME}")
        if str(expected.get("sha256", "")).upper() != _sha256(path):
            raise ValueError(f"{name}: SHA-256 differs from {MANIFEST_NAME}")
        if expected.get("coverage") != inspected["coverage"]:
            raise ValueError(f"{name}: coverage differs from {MANIFEST_NAME}")
        if expected.get("sheets") != inspected["sheets"]:
            raise ValueError(f"{name}: sheet structure differs from {MANIFEST_NAME}")
    if manifest.get("source_files_are_private") is not True:
        raise ValueError(f"{MANIFEST_NAME}: source_files_are_private must be true")
    if manifest.get("original_workbooks_are_not_modified") is not True:
        raise ValueError(f"{MANIFEST_NAME}: original_workbooks_are_not_modified must be true")
    return manifest


def _load_frames(
    paths: dict[str, Path],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    realtime_all: pd.DataFrame | None = None
    day_ahead_all: pd.DataFrame | None = None
    disclosure_all: pd.DataFrame | None = None
    for name in PUBLIC_MARKET_FILES:
        realtime, disclosure = read_workbook(paths[name])
        day_ahead = read_day_ahead_prices(paths[name])
        # Price sheets may contain complete, intentionally blank future days.
        # The existing adapters exclude those days from canonical price data;
        # disclosure rows remain useful and are validated independently.
        if realtime.empty or day_ahead.empty or disclosure.empty:
            raise ValueError(f"{name}: one of the validated market tables is empty")
        realtime_all = merge_identical(realtime_all, realtime, label=f"realtime {name}")
        day_ahead_all = merge_identical(day_ahead_all, day_ahead, label=f"day-ahead {name}")
        disclosure_all = merge_identical(disclosure_all, disclosure, label=f"disclosure {name}")
    manual = read_manual_realtime_prices(paths[PUBLIC_MANUAL_FILE])
    if realtime_all is None or day_ahead_all is None or disclosure_all is None:
        raise ValueError("no market workbooks were loaded")
    return realtime_all, day_ahead_all, disclosure_all, manual


def _validate_content(paths: dict[str, Path]) -> None:
    realtime_all, day_ahead_all, disclosure_all, manual = _load_frames(paths)
    if realtime_all.empty or day_ahead_all.empty or disclosure_all.empty or manual.empty:
        raise ValueError("one of the validated public tables is empty")


def validate_public_package(input_dir: str | Path) -> dict[str, object]:
    """Validate package membership, metadata, schema, hashes, and contracts."""
    input_dir = Path(input_dir).expanduser().resolve()
    paths = _discover(input_dir)
    manifest = _validate_manifest(input_dir, paths)
    _validate_content(paths)
    return manifest


def _runtime_output(runtime_root: Path, relative: str) -> Path:
    runtime = runtime_root.expanduser().resolve()
    output = (runtime / relative).resolve()
    if not output.is_relative_to(runtime):
        raise ValueError("canonical output escaped runtime-root")
    return output


def _save_frame(frame: pd.DataFrame, path: Path, runtime_root: Path) -> str:
    path = path.resolve()
    runtime = runtime_root.expanduser().resolve()
    if not path.is_relative_to(runtime):
        raise ValueError("canonical output must be inside runtime-root")
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_parquet(path)
    return str(path)


def ingest_public_workbooks(
    input_dir: str | Path,
    runtime_root: str | Path,
    *,
    check_only: bool = False,
) -> dict[str, object]:
    """Validate and optionally write canonical parquet outputs."""
    input_dir = Path(input_dir).expanduser().resolve()
    runtime_root = Path(runtime_root).expanduser().resolve()
    paths = _discover(input_dir)
    manifest = _validate_manifest(input_dir, paths)
    if check_only:
        _validate_content(paths)
        return {"check_only": True, "manifest": manifest, "outputs": {}}

    realtime_all, day_ahead_all, disclosure_all, manual = _load_frames(paths)

    output_base = "data/raw/shandong_all_network/SD"
    outputs = {
        "realtime": _save_frame(
            realtime_all,
            _runtime_output(runtime_root, f"{output_base}/realtime_prices_15min.parquet"),
            runtime_root,
        ),
        "day_ahead": _save_frame(
            day_ahead_all,
            _runtime_output(runtime_root, f"{output_base}/day_ahead_prices_15min.parquet"),
            runtime_root,
        ),
        "day_ahead_disclosure": _save_frame(
            disclosure_all,
            _runtime_output(runtime_root, f"{output_base}/day_ahead_disclosure.parquet"),
            runtime_root,
        ),
        "manual_realtime": _save_frame(
            manual,
            _runtime_output(runtime_root, f"{output_base}/manual_realtime_prices_15min.parquet"),
            runtime_root,
        ),
    }
    return {
        "check_only": False,
        "manifest": manifest,
        "outputs": outputs,
        "rows": {
            "realtime": int(len(realtime_all)),
            "day_ahead": int(len(day_ahead_all)),
            "day_ahead_disclosure": int(len(disclosure_all)),
            "manual_realtime": int(len(manual)),
        },
    }


def _resolve(value: str) -> Path:
    return Path(value).expanduser().resolve()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", type=_resolve, required=True)
    parser.add_argument("--runtime-root", type=_resolve, default=Path(".private-runtime"))
    parser.add_argument("--check-only", action="store_true")
    args = parser.parse_args(argv)
    try:
        result = ingest_public_workbooks(args.input_dir, args.runtime_root, check_only=args.check_only)
    except (FileNotFoundError, OSError, ValueError, ImportError) as exc:
        print(json.dumps({"status": "blocked", "error_type": type(exc).__name__, "message": str(exc)}, ensure_ascii=False, indent=2))
        return 2
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
