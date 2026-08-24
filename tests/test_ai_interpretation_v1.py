from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from da_forecast.production.ai_interpretation_v1 import resolve_report_dir, sha256_file, write_ai_interpretation
from da_forecast.production.final_report_v1 import report_directory, write_final_report_artifacts

from test_final_report_artifacts_v1 import _detail, _explanation, _result


def _package(tmp_path: Path):
    result = _result(tmp_path, _detail())
    destination = report_directory(tmp_path, target_date="2026-08-23", run_id=result.run_id)
    artifacts = write_final_report_artifacts(
        tmp_path,
        result=result,
        target_date="2026-08-23",
        as_of="2026-08-22T12:00:00+08:00",
        explanation_payload=_explanation(),
        report_dir=destination,
    )
    report = json.loads(artifacts.report_json_path.read_text(encoding="utf-8"))
    return result, artifacts, report


def _payload(result, artifacts, report) -> dict[str, object]:
    return {
        "schema_version": "ai_interpretation_v1",
        "status": "active",
        "run_id": result.run_id,
        "target_date": "2026-08-23",
        "generated_by": "qilupulse-operation",
        "source": {
            "prediction_checksum": report["prediction_sha256"],
            "whitebox_explanation_checksum": sha256_file(artifacts.explanation_json_path),
            "bundle_parameter_checksum": report["data_coverage"]["bundle_parameter_checksum"],
        },
        "directional_summary": "午间预测低于参考水平，负价风险集中，天气侧证据仅作为描述性支持。",
        "key_explanations": [
            {
                "priority": 1,
                "direction": "downside_support",
                "time_window": "09:00-13:00",
                "statement": "午间预测区间下探，负价概率在该窗口内更高。",
                "evidence_refs": [
                    "period_groups.solar_midday.prediction_summary",
                    "claims:market.recent_price_state",
                ],
                "numeric_evidence": {"max_negative_probability": 0.2},
                "confidence": "descriptive",
                "limitation": "这是历史关联的描述性证据，不是因果证明。",
            }
        ],
        "risk_windows": ["09:00-13:00：关注负价风险。"],
        "data_quality_notes": [],
    }


def test_ai_interpretation_writes_date_partitioned_package_without_changing_final_96(tmp_path: Path) -> None:
    result, artifacts, report = _package(tmp_path)
    artifacts.markdown_path.write_text(
        artifacts.markdown_path.read_text(encoding="utf-8")
        + "\n## 白箱解释\n\n- insufficient_evidence mechanical claim\n",
        encoding="utf-8",
    )
    before = pd.read_excel(artifacts.excel_path, sheet_name="Final_96")

    written = write_ai_interpretation(
        tmp_path,
        run_id=result.run_id,
        target_date="2026-08-23",
        payload=_payload(result, artifacts, report),
    )

    after = pd.read_excel(written.excel_path, sheet_name="Final_96")
    assert written.report_dir == tmp_path / "runs" / "reports" / "2026-08-23" / result.run_id
    assert before.equals(after)
    assert written.json_path.is_file() and written.markdown_path.is_file()
    report_after = json.loads(written.report_json_path.read_text(encoding="utf-8"))
    assert report_after["ai_interpretation_status"] == "active"
    assert report_after["report_revision"] == 1
    workbook = load_workbook(written.excel_path, data_only=True)
    assert "AI_Interpretation" in workbook.sheetnames
    assert "白箱 claim" not in written.report_markdown_path.read_text(encoding="utf-8")
    final_markdown = written.report_markdown_path.read_text(encoding="utf-8")
    assert "insufficient_evidence mechanical claim" not in final_markdown
    assert final_markdown.count("## AI 主要方向性解读") == 1
    assert final_markdown.count("## 文件") == 1

    repeated = write_ai_interpretation(
        tmp_path,
        run_id=result.run_id,
        target_date="2026-08-23",
        payload=_payload(result, artifacts, report),
    )
    assert json.loads(repeated.report_json_path.read_text(encoding="utf-8"))["report_revision"] == 1

    stale = repeated.report_markdown_path.read_text(encoding="utf-8")
    stale = stale.replace("## AI 主要方向性解读", "## 白箱解释\n\n- stale mechanical claim\n\n## AI 主要方向性解读")
    stale += "\n## 文件\n\n- duplicate legacy file list\n"
    repeated.report_markdown_path.write_text(stale.replace("\n", "\r\n"), encoding="utf-8", newline="")
    repaired = write_ai_interpretation(
        tmp_path,
        run_id=result.run_id,
        target_date="2026-08-23",
        payload=_payload(result, artifacts, report),
    )
    repaired_markdown = repaired.report_markdown_path.read_text(encoding="utf-8")
    assert "stale mechanical claim" not in repaired_markdown
    assert repaired_markdown.count("## AI 主要方向性解读") == 1
    assert repaired_markdown.count("## 文件") == 1
    assert json.loads(repaired.report_json_path.read_text(encoding="utf-8"))["report_revision"] == 1


def test_ai_interpretation_rejects_wrong_checksum_without_rewriting_report(tmp_path: Path) -> None:
    result, artifacts, report = _package(tmp_path)
    original = artifacts.report_json_path.read_bytes()
    payload = _payload(result, artifacts, report)
    payload["source"]["prediction_checksum"] = "wrong"

    with pytest.raises(ValueError, match="prediction checksum"):
        write_ai_interpretation(tmp_path, run_id=result.run_id, target_date="2026-08-23", payload=payload)

    assert artifacts.report_json_path.read_bytes() == original
    assert not (artifacts.report_dir / "ai_interpretation.json").exists()


def test_ai_interpretation_rejects_reference_missing_from_whitebox_evidence(tmp_path: Path) -> None:
    result, artifacts, report = _package(tmp_path)
    payload = _payload(result, artifacts, report)
    payload["key_explanations"][0]["evidence_refs"] = ["claims:not-present"]

    with pytest.raises(ValueError, match="references unavailable evidence"):
        write_ai_interpretation(tmp_path, run_id=result.run_id, target_date="2026-08-23", payload=payload)


def test_resolve_report_dir_keeps_legacy_flat_report_readable(tmp_path: Path) -> None:
    legacy = tmp_path / "runs" / "reports" / "legacy-run"
    legacy.mkdir(parents=True)

    assert resolve_report_dir(tmp_path, run_id="legacy-run") == legacy
