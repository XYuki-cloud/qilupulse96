#!/usr/bin/env python3
"""Read-only preflight for one QiluPulse-96 final result package.

The command intentionally prints a small, stable JSON contract so a report
agent does not need to discover directories or infer production readiness from
large CSV/XLSX files.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook


REQUIRED_FINAL_COLUMNS = {
    "period_start",
    "predicted_cny_mwh",
    "p10_cny_mwh",
    "p50_cny_mwh",
    "p90_cny_mwh",
    "negative_probability",
}


def _report_candidates(root: Path, target_date: str | None) -> list[Path]:
    reports = root / "runs" / "reports"
    if not reports.is_dir():
        return []
    if target_date:
        scoped = reports / str(target_date)
        return sorted(scoped.glob("*/final_report.json"), key=lambda p: p.stat().st_mtime, reverse=True) if scoped.is_dir() else []
    partitioned = list(reports.glob("*/**/final_report.json"))
    legacy = list(reports.glob("*/final_report.json"))
    return sorted(partitioned + legacy, key=lambda p: p.stat().st_mtime, reverse=True)


def _resolve_report(root: Path, *, run_id: str | None, target_date: str | None) -> Path:
    candidates = _report_candidates(root, target_date)
    if run_id:
        candidates = [path for path in candidates if path.parent.name == str(run_id)]
    if not candidates:
        selector = f"run_id={run_id}" if run_id else f"target_date={target_date}" if target_date else "latest run"
        raise FileNotFoundError(f"No QiluPulse report package found for {selector}")
    return candidates[0].parent


def _weather_blocked_result(root: Path, *, target_date: str | None) -> dict[str, Any] | None:
    """Expose a weather preflight block even when no prediction package exists."""
    if not target_date:
        return None
    manifest_root = root / "data" / "raw" / "weather_completion"
    if not manifest_root.is_dir():
        return None
    prefix = str(target_date).replace("-", "")
    candidates = sorted(
        manifest_root.glob(f"weather_completion_{prefix}_*.json"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        return None
    manifest_path = candidates[0]
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    if payload.get("status") == "complete":
        return None
    error = str(payload.get("error") or "weather completion is incomplete")
    lowered = error.lower()
    reason = (
        "target weather snapshot missing"
        if "target forecast snapshot missing" in lowered
        else "target weather snapshot timestamp mismatch"
        if "issued_at mismatch" in lowered
        else "weather completion blocked"
    )
    return {
        "status": "blocked",
        "reason": reason,
        "target_date": str(target_date),
        "paths": {"weather_completion_manifest": str(manifest_path)},
        "errors": [error],
    }


def _read_csv_shape(path: Path) -> tuple[int, set[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fields = set(reader.fieldnames or [])
        rows = sum(1 for _ in reader)
    return rows, fields


def inspect_result(root: str | Path, *, run_id: str | None = None, target_date: str | None = None) -> dict[str, Any]:
    root = Path(root).expanduser().resolve()
    try:
        report_dir = _resolve_report(root, run_id=run_id, target_date=target_date)
    except FileNotFoundError:
        weather_block = _weather_blocked_result(root, target_date=target_date)
        if weather_block is not None:
            return weather_block
        raise
    report_path = report_dir / "final_report.json"
    report = json.loads(report_path.read_text(encoding="utf-8"))
    resolved_run_id = str(report.get("run_id") or report_dir.name)
    resolved_target_date = str(report.get("target_date") or "")
    prediction_dir = root / "runs" / "predictions" / resolved_run_id
    metadata_path = prediction_dir / "run_metadata.json"
    detail_path = prediction_dir / "prediction_detail.csv"
    whitebox_path = report_dir / "whitebox_explanation.json"
    excel_path = report_dir / "final_prediction.xlsx"
    errors: list[str] = []

    if target_date and resolved_target_date != str(target_date):
        errors.append(f"target_date mismatch: expected {target_date}, got {resolved_target_date}")
    if run_id and resolved_run_id != str(run_id):
        errors.append(f"run_id mismatch: expected {run_id}, got {resolved_run_id}")
    metadata: dict[str, Any] = {}
    if metadata_path.is_file():
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    else:
        errors.append(f"missing run metadata: {metadata_path}")
    if not whitebox_path.is_file():
        errors.append(f"missing white-box evidence: {whitebox_path}")
    if not excel_path.is_file():
        errors.append(f"missing final workbook: {excel_path}")

    detail_rows = 0
    detail_columns: set[str] = set()
    if detail_path.is_file():
        detail_rows, detail_columns = _read_csv_shape(detail_path)
        if detail_rows != 96:
            errors.append(f"prediction_detail must have 96 rows, got {detail_rows}")
    else:
        errors.append(f"missing prediction detail: {detail_path}")

    final_rows = 0
    final_columns: set[str] = set()
    if excel_path.is_file():
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            if "Final_96" not in workbook.sheetnames:
                errors.append("final workbook is missing Final_96 sheet")
            else:
                sheet = workbook["Final_96"]
                values = list(sheet.iter_rows(values_only=True))
                final_columns = {str(value) for value in (values[0] if values else ()) if value is not None}
                final_rows = max(0, len(values) - 1)
                if final_rows != 96:
                    errors.append(f"Final_96 must have 96 rows, got {final_rows}")
                missing = sorted(REQUIRED_FINAL_COLUMNS - final_columns)
                if missing:
                    errors.append(f"Final_96 missing columns: {', '.join(missing)}")
        finally:
            workbook.close()

    coverage = report.get("data_coverage") or {}
    report_checksum = coverage.get("bundle_parameter_checksum")
    metadata_checksum = metadata.get("parameter_checksum")
    if report_checksum and metadata_checksum and str(report_checksum) != str(metadata_checksum):
        errors.append("bundle parameter checksum mismatch between report and run metadata")
    if report.get("calibration_status") != "active":
        errors.append(f"calibration_status is {report.get('calibration_status')!r}; production report requires active")
    if not report.get("prediction_sha256"):
        errors.append("report is missing prediction_sha256")

    return {
        "status": "ready" if not errors else "blocked",
        "run_id": resolved_run_id,
        "target_date": resolved_target_date,
        "as_of": report.get("as_of"),
        "publish_status": report.get("publish_status"),
        "calibration_status": report.get("calibration_status"),
        "calibration_history_days": report.get("calibration_history_days"),
        "ai_interpretation_status": report.get("ai_interpretation_status", "pending"),
        "row_count": report.get("row_count"),
        "detail_rows": detail_rows,
        "final_96_rows": final_rows,
        "bundle_parameter_checksum": report_checksum,
        "realtime_cutoff": coverage.get("realtime_cutoff"),
        "paths": {
            "report_dir": str(report_dir),
            "final_report_json": str(report_path),
            "final_report_markdown": str(report_dir / "final_report.md"),
            "final_prediction_png": str(report_dir / "final_prediction.png"),
            "final_prediction_xlsx": str(excel_path),
            "whitebox_explanation_json": str(whitebox_path),
            "ai_interpretation_json": str(report_dir / "ai_interpretation.json"),
            "ai_interpretation_markdown": str(report_dir / "ai_interpretation.md"),
            "run_metadata": str(metadata_path),
            "prediction_detail": str(detail_path),
        },
        "errors": errors,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--target-date", help="Use the newest package under this YYYY-MM-DD directory")
    parser.add_argument("--run-id", help="Read one exact run package")
    args = parser.parse_args(argv)
    try:
        result = inspect_result(args.root, run_id=args.run_id, target_date=args.target_date)
    except (FileNotFoundError, OSError, json.JSONDecodeError, ValueError) as exc:
        result = {"status": "blocked", "errors": [f"{type(exc).__name__}: {exc}"]}
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 2
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["status"] == "ready" else 2


if __name__ == "__main__":
    raise SystemExit(main())
