from __future__ import annotations

import json
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

from da_forecast.production.final_report_v1 import write_final_report_artifacts


def _detail(target_date: str = "2026-08-23") -> pd.DataFrame:
    slots = pd.date_range(target_date, periods=96, freq="15min")
    return pd.DataFrame(
        {
            "market_date": slots.strftime("%Y-%m-%d"),
            "period_start": slots.strftime("%H:%M"),
            "predicted_cny_mwh": [300.0 + index for index in range(96)],
            "negative_probability": [index / 9_600 for index in range(96)],
            "p10_cny_mwh": [200.0 + index for index in range(96)],
            "p50_cny_mwh": [300.0 + index for index in range(96)],
            "p90_cny_mwh": [400.0 + index for index in range(96)],
            "raw_predicted_cny_mwh": [100.0 + index for index in range(96)],
            "raw_p10_cny_mwh": [50.0 + index for index in range(96)],
            "raw_p50_cny_mwh": [100.0 + index for index in range(96)],
            "raw_p90_cny_mwh": [150.0 + index for index in range(96)],
            "bias_correction_cny_mwh": [200.0] * 96,
            "interval_lower_expansion_cny_mwh": [0.0] * 96,
            "interval_upper_expansion_cny_mwh": [50.0] * 96,
            "bias_status": ["active"] * 96,
            "interval_status": ["active"] * 96,
        }
    )


def _result(root, detail: pd.DataFrame) -> SimpleNamespace:
    run_dir = root / "runs" / "predictions" / "run-1"
    run_dir.mkdir(parents=True)
    detail_path = run_dir / "prediction_detail.csv"
    detail.to_csv(detail_path, index=False)
    return SimpleNamespace(
        run_id="run-1",
        publish_status="draft",
        detail_path=detail_path,
        metadata={
            "calibration": {"calibration_status": "active", "calibration_history_days": 56},
            "realtime_cutoff": "2026-08-22T10:45:00+08:00",
            "weather_source_hash": "weather-hash",
            "weather_source_counts": {"observed": 10, "target_forecast": 5},
            "parameter_checksum": "bundle-checksum",
        },
    )


def _explanation() -> dict[str, object]:
    return {
        "explanation_version": "whitebox_v1.1",
        "market_state": {"price_state": "接近常态", "z_price": 0.1},
        "claims": [
            {
                "claim_id": "market.recent_price_state",
                "claim_type": "market_state",
                "period_group": None,
                "statement": "这是可观测市场状态描述，不是对未来价格的确定承诺。",
                "confidence_level": "descriptive",
                "reference_window": "recent_7d_vs_recent_90d",
                "effect_estimate": 0.1,
            }
        ],
        "period_groups": {
            "solar_midday": {
                "slot_start": 36,
                "slot_end": 59,
                "reference_status": "calendar_matched",
                "prediction_summary": {"mean_predicted_cny_mwh": 320.0, "max_negative_probability": 0.2},
            }
        },
    }


def test_write_final_artifacts_uses_calibrated_96_slot_result_and_groups_every_file(tmp_path) -> None:
    result = _result(tmp_path, _detail())

    artifacts = write_final_report_artifacts(
        tmp_path,
        result=result,
        target_date="2026-08-23",
        as_of="2026-08-22T12:00:00+08:00",
        explanation_payload=_explanation(),
    )

    report_dir = tmp_path / "runs" / "reports" / "run-1"
    expected = {
        "final_prediction.png",
        "final_prediction.xlsx",
        "final_report.md",
        "final_report.json",
        "whitebox_explanation.json",
    }
    assert expected <= {path.name for path in report_dir.iterdir()}
    assert all((report_dir / name).stat().st_size > 0 for name in expected)

    workbook = load_workbook(artifacts.excel_path, data_only=True)
    assert workbook.sheetnames == ["Final_96", "Hourly_24", "Explanation", "Audit"]
    final = workbook["Final_96"]
    assert final.max_row == 97
    columns = {cell.value: cell.column for cell in final[1]}
    assert final.cell(row=2, column=columns["predicted_cny_mwh"]).value == 300.0
    assert final.cell(row=2, column=columns["raw_predicted_cny_mwh"]).value == 100.0

    payload = json.loads(artifacts.report_json_path.read_text(encoding="utf-8"))
    assert payload["row_count"] == 96
    assert payload["calibration_status"] == "active"
    assert payload["explanation_status"] == "active"
    assert "未经后处理的 raw 结果不作为最终预测" in artifacts.markdown_path.read_text(encoding="utf-8")
