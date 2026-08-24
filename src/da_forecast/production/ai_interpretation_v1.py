"""Validate and atomically write AI directionality interpretations.

The production model and white-box explainer remain deterministic.  This module
only persists a separately generated interpretation after checking that it was
written for the exact prediction package it describes.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
import json
from pathlib import Path
import os
import tempfile
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


SCHEMA_VERSION = "ai_interpretation_v1"
VALID_DIRECTIONS = {"upside_support", "downside_support", "mixed", "neutral"}
VALID_CONFIDENCE = {"supported", "descriptive", "uncertain"}


@dataclass(frozen=True)
class AIInterpretationArtifacts:
    report_dir: Path
    json_path: Path
    markdown_path: Path
    report_json_path: Path
    report_markdown_path: Path
    excel_path: Path
    status: str


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_report_dir(root: str | Path, *, run_id: str, target_date: str | None = None) -> Path:
    """Prefer date-partitioned reports, then read the legacy flat layout."""
    root = Path(root)
    reports = root / "runs" / "reports"
    if target_date:
        candidate = reports / str(target_date) / str(run_id)
        if candidate.is_dir():
            return candidate
    matches = sorted(reports.glob(f"*/{run_id}")) if reports.is_dir() else []
    if matches:
        return matches[0]
    legacy = reports / str(run_id)
    if legacy.is_dir():
        return legacy
    raise FileNotFoundError(f"Unknown report package: {run_id}")


def validate_interpretation(
    payload: dict[str, Any],
    *,
    report: dict[str, Any],
    whitebox_checksum: str,
    whitebox_payload: dict[str, Any] | None,
) -> dict[str, Any]:
    """Validate a skill-produced payload without evaluating its prose."""
    if not isinstance(payload, dict):
        raise ValueError("AI interpretation must be a JSON object")
    if payload.get("schema_version") != SCHEMA_VERSION:
        raise ValueError(f"AI interpretation schema_version must be {SCHEMA_VERSION}")
    if payload.get("status") not in {"active", "unavailable"}:
        raise ValueError("AI interpretation status must be active or unavailable")
    if str(payload.get("run_id")) != str(report.get("run_id")):
        raise ValueError("AI interpretation run_id does not match the report")
    if str(payload.get("target_date")) != str(report.get("target_date")):
        raise ValueError("AI interpretation target_date does not match the report")
    source = payload.get("source")
    if not isinstance(source, dict):
        raise ValueError("AI interpretation source is required")
    expected_prediction = str(report.get("prediction_sha256"))
    expected_bundle = str(report.get("data_coverage", {}).get("bundle_parameter_checksum"))
    if str(source.get("prediction_checksum")) != expected_prediction:
        raise ValueError("AI interpretation prediction checksum does not match the report")
    if str(source.get("whitebox_explanation_checksum")) != whitebox_checksum:
        raise ValueError("AI interpretation whitebox checksum does not match the evidence file")
    if str(source.get("bundle_parameter_checksum")) != expected_bundle:
        raise ValueError("AI interpretation bundle checksum does not match the report")
    if not str(payload.get("generated_by", "")).strip():
        raise ValueError("AI interpretation generated_by is required")
    summary = str(payload.get("directional_summary", "")).strip()
    if not summary:
        raise ValueError("AI interpretation directional_summary is required")
    explanations = payload.get("key_explanations", [])
    if not isinstance(explanations, list):
        raise ValueError("AI interpretation key_explanations must be a list")
    for index, item in enumerate(explanations):
        if not isinstance(item, dict):
            raise ValueError(f"key_explanations[{index}] must be an object")
        if item.get("direction") not in VALID_DIRECTIONS:
            raise ValueError(f"key_explanations[{index}] has an invalid direction")
        if item.get("confidence") not in VALID_CONFIDENCE:
            raise ValueError(f"key_explanations[{index}] has an invalid confidence")
        if not str(item.get("statement", "")).strip():
            raise ValueError(f"key_explanations[{index}] statement is required")
        refs = item.get("evidence_refs")
        if not isinstance(refs, list) or not refs or not all(isinstance(ref, str) and ref.strip() for ref in refs):
            raise ValueError(f"key_explanations[{index}] requires non-empty evidence_refs")
        for ref in refs:
            if not _evidence_ref_exists(ref, report=report, whitebox_payload=whitebox_payload or {}):
                raise ValueError(f"key_explanations[{index}] references unavailable evidence: {ref}")
        if not str(item.get("time_window", "")).strip():
            raise ValueError(f"key_explanations[{index}] time_window is required")
    for field in ("risk_windows", "data_quality_notes"):
        if not isinstance(payload.get(field, []), list):
            raise ValueError(f"AI interpretation {field} must be a list")
    normalized = json.loads(json.dumps(payload, ensure_ascii=False, default=str))
    normalized["generated_at"] = normalized.get("generated_at") or datetime.now(timezone.utc).isoformat()
    return normalized


def write_ai_interpretation(
    root: str | Path,
    *,
    run_id: str,
    payload: dict[str, Any],
    target_date: str | None = None,
) -> AIInterpretationArtifacts:
    """Write AI interpretation files and update the final report atomically."""
    root = Path(root)
    report_dir = resolve_report_dir(root, run_id=run_id, target_date=target_date)
    report_json_path = report_dir / "final_report.json"
    report_markdown_path = report_dir / "final_report.md"
    excel_path = report_dir / "final_prediction.xlsx"
    whitebox_path = report_dir / "whitebox_explanation.json"
    if not report_json_path.is_file() or not report_markdown_path.is_file() or not excel_path.is_file() or not whitebox_path.is_file():
        raise FileNotFoundError(f"Incomplete final report package: {report_dir}")
    report = json.loads(report_json_path.read_text(encoding="utf-8"))
    whitebox_checksum = sha256_file(whitebox_path)
    whitebox_document = json.loads(whitebox_path.read_text(encoding="utf-8"))
    whitebox_payload = whitebox_document.get("payload", whitebox_document)
    ai_json_path = report_dir / "ai_interpretation.json"
    existing = json.loads(ai_json_path.read_text(encoding="utf-8")) if ai_json_path.is_file() else None
    if isinstance(existing, dict) and "generated_at" not in payload:
        payload = {**payload, "generated_at": existing.get("generated_at")}
    normalized = validate_interpretation(
        payload,
        report=report,
        whitebox_checksum=whitebox_checksum,
        whitebox_payload=whitebox_payload if isinstance(whitebox_payload, dict) else None,
    )
    normalized["source"]["whitebox_explanation_checksum"] = whitebox_checksum
    ai_markdown_path = report_dir / "ai_interpretation.md"
    normalized["artifact_paths"] = {
        "ai_interpretation_json": ai_json_path.name,
        "ai_interpretation_markdown": ai_markdown_path.name,
    }
    ai_markdown = render_ai_markdown(normalized)
    existing_markdown = report_markdown_path.read_text(encoding="utf-8")
    markdown = _update_final_markdown(existing_markdown, normalized)
    same_interpretation = existing == normalized
    report_is_current = (
        report.get("ai_interpretation_status") == normalized["status"]
        and report.get("ai_interpretation", {}).get("directional_summary") == normalized["directional_summary"]
    )
    workbook_is_current = _workbook_has_ai_interpretation(excel_path)
    if (
        same_interpretation
        and report_is_current
        and workbook_is_current
        and existing_markdown == markdown
        and ai_markdown_path.is_file()
        and ai_markdown_path.read_text(encoding="utf-8") == ai_markdown
    ):
        return AIInterpretationArtifacts(
            report_dir=report_dir,
            json_path=ai_json_path,
            markdown_path=ai_markdown_path,
            report_json_path=report_json_path,
            report_markdown_path=report_markdown_path,
            excel_path=excel_path,
            status=str(normalized["status"]),
        )
    report["ai_interpretation_status"] = normalized["status"]
    report["ai_interpretation"] = {
        "status": normalized["status"],
        "directional_summary": normalized["directional_summary"],
        "key_explanations": normalized.get("key_explanations", []),
        "risk_windows": normalized.get("risk_windows", []),
        "data_quality_notes": normalized.get("data_quality_notes", []),
        "generated_at": normalized.get("generated_at"),
    }
    report.setdefault("artifact_paths", {})["ai_interpretation_json"] = ai_json_path.name
    report.setdefault("artifact_paths", {})["ai_interpretation_markdown"] = ai_markdown_path.name
    if not same_interpretation:
        report["report_revision"] = int(report.get("report_revision", 0)) + 1

    _atomic_write_json(ai_json_path, normalized)
    _atomic_write_text(ai_markdown_path, ai_markdown)
    _atomic_write_json(report_json_path, report)
    _atomic_write_text(report_markdown_path, markdown)
    _update_workbook(excel_path, normalized)
    _update_run_metadata(
        root,
        run_id=run_id,
        report_dir=report_dir,
        normalized=normalized,
        increment_revision=not same_interpretation,
    )
    return AIInterpretationArtifacts(
        report_dir=report_dir,
        json_path=ai_json_path,
        markdown_path=ai_markdown_path,
        report_json_path=report_json_path,
        report_markdown_path=report_markdown_path,
        excel_path=excel_path,
        status=str(normalized["status"]),
    )


def render_ai_markdown(payload: dict[str, Any]) -> str:
    lines = [
        f"# AI 方向性解读（{payload['target_date']}）",
        "",
        f"- 运行编号：`{payload['run_id']}`",
        f"- 状态：`{payload['status']}`",
        f"- 生成来源：`{payload['generated_by']}`",
        "",
        "## 主要方向",
        "",
        str(payload["directional_summary"]),
    ]
    explanations = payload.get("key_explanations", [])
    if explanations:
        lines.extend(["", "## 关键解释", ""])
        for item in explanations:
            lines.append(
                f"- **{item['time_window']} / {item['direction']} / {item['confidence']}**：{item['statement']}"
            )
            lines.append(f"  - 证据：`{', '.join(item['evidence_refs'])}`")
            if item.get("limitation"):
                lines.append(f"  - 限制：{item['limitation']}")
    if payload.get("risk_windows"):
        lines.extend(["", "## 风险窗口", ""])
        lines.extend(f"- {item}" for item in payload["risk_windows"])
    if payload.get("data_quality_notes"):
        lines.extend(["", "## 相关数据限制", ""])
        lines.extend(f"- {item}" for item in payload["data_quality_notes"])
    lines.extend(
        [
            "",
            "本解读由 AI 基于本次预测、白箱证据和数据覆盖生成；它是可复核的描述性分析，不是电力市场因果证明，也不改变模型预测数值。",
            "",
        ]
    )
    return "\n".join(lines)


def _evidence_ref_exists(ref: str, *, report: dict[str, Any], whitebox_payload: dict[str, Any]) -> bool:
    """Allow only references that resolve inside this immutable result package."""
    if ref.startswith("claims:"):
        claim_id = ref.removeprefix("claims:")
        return any(str(claim.get("claim_id")) == claim_id for claim in whitebox_payload.get("claims", []))
    if ref.startswith("prediction."):
        return ref.removeprefix("prediction.") in {
            "mean_predicted_cny_mwh", "peak", "valley", "max_negative_probability", "final_96",
        }
    if ref.startswith("data_coverage."):
        return _has_path(report.get("data_coverage", {}), ref.removeprefix("data_coverage.").split("."))
    for prefix in ("market_state.", "calendar.", "period_groups."):
        if ref.startswith(prefix):
            return _has_path(whitebox_payload.get(prefix.removesuffix("."), {}), ref.removeprefix(prefix).split("."))
    return False


def _has_path(value: Any, parts: list[str]) -> bool:
    current = value
    for part in parts:
        if not isinstance(current, dict) or part not in current:
            return False
        current = current[part]
    return True


def _update_final_markdown(existing: str, payload: dict[str, Any]) -> str:
    start = "## AI 主要方向性解读"
    end = "## 文件"
    # Older packages embedded every white-box claim in the main report.  Keep
    # the deterministic audit JSON, but remove that mechanical section from
    # the user-facing report before inserting the filtered AI interpretation.
    cleaned = _normalize_markdown(existing)
    cleaned = _remove_markdown_section(cleaned, "## 白箱解释")
    cleaned = _remove_markdown_section(cleaned, start)
    cleaned = _keep_first_markdown_section(cleaned, end)
    before = cleaned.split(end, 1)[0] if end in cleaned else cleaned
    suffix = cleaned.split(end, 1)[1] if end in cleaned else ""
    section = [start, "", f"- 状态：`{payload['status']}`。", "", str(payload["directional_summary"]), ""]
    explanations = payload.get("key_explanations", [])
    for item in explanations:
        section.append(f"- **{item['time_window']} / {item['direction']} / {item['confidence']}**：{item['statement']}")
        section.append(f"  - 证据：`{', '.join(item['evidence_refs'])}`")
        if item.get("limitation"):
            section.append(f"  - 限制：{item['limitation']}")
    if payload.get("risk_windows"):
        section.extend(["", "### 风险窗口", "", *[f"- {item}" for item in payload["risk_windows"]]])
    if payload.get("data_quality_notes"):
        section.extend(["", "### 相关数据限制", "", *[f"- {item}" for item in payload["data_quality_notes"]]])
    section.extend(
        [
            "",
            "以上方向性解读由 AI 基于本次预测、白箱证据和数据覆盖生成；它是可复核的描述性分析，不是电力市场因果证明，也不改变模型预测数值。",
            "",
        ]
    )
    return before.rstrip() + "\n\n" + "\n".join(section) + "\n" + (end + suffix if end in cleaned else "")


def _normalize_markdown(text: str) -> str:
    return text.replace("\r\n", "\n").replace("\r", "\n")


def _remove_markdown_section(text: str, heading: str) -> str:
    """Remove one level-2 section, up to the next level-2 heading."""
    lines = text.splitlines()
    kept: list[str] = []
    index = 0
    while index < len(lines):
        if lines[index] != heading:
            kept.append(lines[index])
            index += 1
            continue
        index += 1
        while index < len(lines) and not lines[index].startswith("## "):
            index += 1
    return "\n".join(kept).rstrip() + "\n"


def _keep_first_markdown_section(text: str, heading: str) -> str:
    """Drop duplicate copies of a level-2 section while preserving the first."""
    lines = text.splitlines()
    kept: list[str] = []
    seen = False
    index = 0
    while index < len(lines):
        if lines[index] != heading or not seen:
            if lines[index] == heading:
                seen = True
            kept.append(lines[index])
            index += 1
            continue
        index += 1
        while index < len(lines) and not lines[index].startswith("## "):
            index += 1
    return "\n".join(kept).rstrip() + "\n"


def _workbook_has_ai_interpretation(path: Path) -> bool:
    workbook = load_workbook(path, read_only=True)
    try:
        return "AI_Interpretation" in workbook.sheetnames
    finally:
        workbook.close()


def _update_workbook(path: Path, payload: dict[str, Any]) -> None:
    workbook = load_workbook(path)
    if "AI_Interpretation" in workbook.sheetnames:
        del workbook["AI_Interpretation"]
    sheet = workbook.create_sheet("AI_Interpretation")
    headers = ["priority", "direction", "time_window", "statement", "confidence", "evidence_refs", "numeric_evidence", "limitation"]
    sheet.append(headers)
    for item in payload.get("key_explanations", []):
        sheet.append([
            item.get("priority"), item.get("direction"), item.get("time_window"), item.get("statement"),
            item.get("confidence"), ", ".join(item.get("evidence_refs", [])),
            json.dumps(item.get("numeric_evidence", {}), ensure_ascii=False, default=str), item.get("limitation"),
        ])
    if not payload.get("key_explanations"):
        sheet.append(["status", payload.get("status"), "", payload.get("directional_summary"), "", "", "", ""])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column_letter].width = min(60, max(14, max(len(str(cell.value or "")) for cell in column_cells) + 2))
    _atomic_save_workbook(path, workbook)


def _update_run_metadata(
    root: Path,
    *,
    run_id: str,
    report_dir: Path,
    normalized: dict[str, Any],
    increment_revision: bool,
) -> None:
    metadata_path = root / "runs" / "predictions" / run_id / "run_metadata.json"
    if not metadata_path.is_file():
        return
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    metadata.update(
        {
            "ai_interpretation_status": normalized["status"],
            "ai_interpretation_json_path": str((report_dir / "ai_interpretation.json").relative_to(root)),
            "ai_interpretation_markdown_path": str((report_dir / "ai_interpretation.md").relative_to(root)),
        }
    )
    if increment_revision:
        metadata["report_revision"] = int(metadata.get("report_revision", 0)) + 1
    _atomic_write_json(metadata_path, metadata)


def _atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="") as handle:
            handle.write(text)
        os.replace(temp_name, path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)


def _atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    _atomic_write_text(path, json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n")


def _atomic_save_workbook(path: Path, workbook) -> None:
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".xlsx", dir=path.parent)
    os.close(fd)
    try:
        workbook.save(temp_name)
        os.replace(temp_name, path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)
